# 20170727
# Flora Tsai

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import getpass
import sys
import openpyxl
from openpyxl import Workbook
import argparse
import time

'''
Global Variables
'''
# Start url
url = 'https://example.splunk.com/en-US/app/search/reports'

# Columns in spreadsheet
col_index = {'Title': 1, 'Query': 2, 'Time': 3}

# Spreadsheet start row
start_row = 2

# Output filename
filename = 'Reports.xlsx'

# Set web browser
# driver = webdriver.Firefox()
driver = webdriver.PhantomJS()

'''
Code start
'''
# Create a new spreadsheet
wb = Workbook()
wb_sheet = wb.create_sheet(title='Sheet1')
wb_sheet.cell(row=1, column=col_index['Title']).value = 'Report Title'
wb_sheet.cell(row=1, column=col_index['Query']).value = 'Search Query'
wb_sheet.cell(row=1, column=col_index['Time']).value = 'Time Range'

# Arg parser
parser = argparse.ArgumentParser()
parser.add_argument('-f', '--file', help='Set customized filename')
args = parser.parse_args()
if args.file:
	filename = args.file

# Connect to url with login credential
driver.get(url)
user = input('Enter username: ')
passwd = getpass.getpass()
username = driver.find_element_by_id('username')
username.send_keys(user)
password = driver.find_element_by_id('password')
password.send_keys(passwd)
btn = driver.find_element_by_id('login-submit')
btn.click()

'''
# Two-factor authentication with duo app
try:
	WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.NAME, 'send')))
except TimeoutException:
	print('Two-factor authentication failed')
	driver.quit()
	sys.exit()

# btn2 = driver.find_element_by_name('send')
# btn2.click()


# Redirecting
try:
	WebDriverWait(driver, 40).until(EC.title_contains('Reports'))
except TimeoutException:
	print('Please approve the request on your device within 40 seconds')
	driver.quit()
	sys.exit()
'''

try:
	WebDriverWait(driver, 20).until(EC.title_contains('Reports'))
	WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.NAME, 'display.prefs.aclFilter')))
except TimeoutException:
	print('Tab error')
	driver.quit()
	sys.exit()

# All reports
tab = driver.find_element_by_xpath('//button[@name="display.prefs.aclFilter"][@data-value="none"]')

# Yours reports
# tab = driver.find_element_by_xpath('//button[@name="display.prefs.aclFilter"][@data-value="owner"]')

tab.click()
# Make sure all reports are loaded
time.sleep(3)

try:
	WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'openInLink')))
except TimeoutException:
	print('Reports cannot be loaded')
	driver.quit()
	sys.exit()


elem_list = driver.find_elements_by_xpath('//a[@class="openInLink"]')

# Report urls
url_list = []

# Debug purpose
datasets = []

for u in elem_list:
	if u.text == 'Open in Search':
		link = u.get_attribute('href')
		url_list.append(link)
		
for i in url_list:
	driver.get(i)
	try:
		WebDriverWait(driver, 20).until(EC.title_contains('Search'))
		WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//h2[@class="search-name section-title search-title-searchname"]')))
		WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//textarea[@name="q"]')))
		WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//span[@class="time-label"]')))
	except TimeoutException:
		print('Search %s cannot be loaded' %i)
		continue

	# Debug purpose
	mydict = {}
	mydict['Title'] = driver.find_element_by_xpath('//h2[@class="search-name section-title search-title-searchname"]').get_attribute('title')
	mydict['Query'] = driver.find_element_by_xpath('//textarea[@name="q"]').text
	mydict['Time Range'] = driver.find_element_by_xpath('//span[@class="time-label"]').text
	datasets.append(mydict)
	
	# Write to spreadsheet
	wb_sheet.cell(row=start_row, column=col_index['Title']).value = driver.find_element_by_xpath('//h2[@class="search-name section-title search-title-searchname"]').get_attribute('title')
	wb_sheet.cell(row=start_row, column=col_index['Query']).value = driver.find_element_by_xpath('//textarea[@name="q"]').text
	wb_sheet.cell(row=start_row, column=col_index['Time']).value = driver.find_element_by_xpath('//span[@class="time-label"]').text
	start_row += 1



print(len(url_list))
print(len(datasets))
print(datasets)

# Save spreadsheet
wb.save(filename)
print('Saved as ' + filename)

# Close the browser
driver.quit()

